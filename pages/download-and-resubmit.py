

import dash
from dash import dcc, html,callback
from dash.dependencies import Input, Output, State

import plotly.express as px
import dash_bootstrap_components as dbc

from dash.exceptions import PreventUpdate

import numpy as np
import pandas as pd

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import openpyxl

import base64
import io
from pprint import pprint

dash.register_page(__name__, path='/download-and-resubmit')

#we call the keys here "archetypes"
FORM_HEADER_DICT={
    'tissue':['species','organ','sex','height','heightUnit','weight','weightUnit','age','ageUnit','mass','massUnit','otherInclusionCriteria','otherExclusionCriteria'],
    'fluid':['species','organ','sex','height','heightUnit','weight','weightUnit','age','ageUnit','volume','volumeUnit','otherInclusionCriteria','otherExclusionCriteria'],
    'cells':['species','cellLine','cellCount','otherInclusionCriteria','otherExclusionCriteria'],
    'raw_material':['material','mass','massUnit','volume','volumeUnit','otherInclusionCriteria','otherExclusionCriteria'],
    'genetic':['gene'],
    'longitudinal':['zeroTimeEvent','time','timeUnit'],
    'intervention':['drugName','drugDoseMagnitude','drugDoseUnit','diet','exercise'],
    'effect':['disease'],
}

#obtain the reverse of the FORM_HEADER_DICT, eg {'species:['tissue','fluid','cells']}
FORM_HEADER_DICT_REVERSE={element:set() for element in sum(FORM_HEADER_DICT.values(),[])}
for temp_archetype in FORM_HEADER_DICT.keys():
    for temp_header in FORM_HEADER_DICT[temp_archetype]:
        FORM_HEADER_DICT_REVERSE[temp_header].add(temp_archetype)

SPLIT_CHAR='~'

COLOR_LIST=['red','orange','yellow','green','lime','sky','khaki','red','orange','yellow','green','lime','sky','khaki','red','orange','yellow','green','lime','sky','khaki']
    

def generate_form_headers(selected_archetypes):
    '''
    from the selected archetypes ('tissue', 'genetic', etc.)
    create the total set of metadata headers. order matters 
    '''
    total_headers=[]
    for temp_header in selected_archetypes:
        for temp_element in FORM_HEADER_DICT[temp_header]:
            if temp_element not in total_headers:
                total_headers.append(temp_element)
    
    return total_headers

def split_columns_if_delimited(temp_dataframe):

    #for each column
    #split it
    #delete the orignal
    #append the new ones
    #much easier to conserve the order than to reorder
    #do in parallel with temp_dataframe_2
    new_dataframe_list=list()
    #new_dataframe_list_2=list()
    for temp_column in temp_dataframe.columns:
        if temp_dataframe[temp_column].dtype==object:
            #num_cols_after_split=len(temp_dataframe[temp_column].str.split(SPLIT_CHAR,expand=True))
            temp_expanded_columns=temp_dataframe[temp_column].str.split(SPLIT_CHAR,expand=True).add_prefix(temp_column+'.')
            #temp_expanded_columns_2=temp_dataframe_2
        else:
            temp_expanded_columns=temp_dataframe[temp_column]
        new_dataframe_list.append(temp_expanded_columns)

    output_dataframe=pd.concat(new_dataframe_list,axis='columns')

    output_dataframe.fillna(value=np.nan,inplace=True)

    return output_dataframe

layout = html.Div(
    children=[
        dcc.Download(id="download_form"),
        html.Br(),
        html.Br(),
        dbc.Row(
            children=[
                dbc.Col(width=3),
                dbc.Col(
                    children=[
                        html.H3('Sample Types'),
                        html.Br(),
                        dbc.Checklist(
                            options=[
                                {"label": "Tissue (lung, heart, etc.)", "value": 'tissue'},
                                {"label": "Biofluids (plasma, urine, etc.)", "value": 'fluid'},
                                {"label": "Cells (culture, organoid, etc.)", "value": 'cells'},
                                {"label": "Raw Material (soil, water, gas, etc.)", "value": 'raw_material'},
                            ],
                            id="sample_checklist",
                        ),
                    ],
                    width=4
                ),
                dbc.Col(
                    children=[
                        html.H3('Study Types'),
                        html.Br(),
                        dbc.Checklist(
                            options=[
                                {"label": "Genetic (knockout, CRISPR, MIR, etc.)", "value": 'genetic'},
                                {"label": "Time Series (longitudinal)", "value": 'longitudinal'},
                                {"label": "Intervention (drug, diet, exercise, etc.)", "value": 'intervention'},
                                {"label": "Effect (disease, etc.)", "value": 'effect'},
                            ],
                            id="study_checklist",
                        ),
                    ],
                    width=4
                ),
                dbc.Col(width=1)
            ]
        ),
        html.Br(),
        html.Br(),
        html.Br(),
        html.Br(),
        dbc.Row(
            children=[
                dbc.Col(width=5),
                dbc.Col(
                    children=[
                        html.Div(
                            dbc.Button(
                                'Download Form',
                                id='button_form',
                            ),
                            className="d-grid gap-2 col-6 mx-auto",
                        ),
                    ],
                    width=2
                ),
                dbc.Col(width=5)
            ]
        ),
        html.Br(),
        html.Br(),
        dbc.Row(
            children=[
                dbc.Col(width=4),
                dbc.Col(
                    children=[
                        dcc.Upload(
                            id='upload_form',
                            children=html.Div([
                                'Drag and Drop or Select Files',
                            ]),
                            style={
                                'width': '100%',
                                'height': '60px',
                                'lineHeight': '60px',
                                'borderWidth': '1px',
                                'borderStyle': 'dashed',
                                'borderRadius': '5px',
                                'textAlign': 'center',
                                'margin': '10px'
                            },
                        ),
                    ],
                    width=4
                ),
                dbc.Col(width=4)
            ]
        ),
        dbc.Row(
            children=[
                dbc.Col(width=5),
                dbc.Col(
                    children=[
                        html.Div(id='div_filename')
                    ],
                    width=2
                ),
                dbc.Col(width=5)
            ]
        ),
        html.Br(),
        html.Br(),
        dbc.Row(
            children=[
                dbc.Col(width=5),
                dbc.Col(
                    children=[
                        
                        dcc.Link(
                            children=[                        
                                html.Div(
                                    dbc.Button(
                                        'Process Form',
                                        id='button_form',
                                    ),
                                    className="d-grid gap-2 col-6 mx-auto",
                                ),
                            ],
                            href='/curate-and-download',
                        )
                    ],
                    width=2
                ),
                dbc.Col(width=5)
            ]
        ),
    ],
)


def generate_header_colors(selected_archetypes,total_headers):
    '''
    this generates the dictionaries that are used to color the top row in the excel file
    we use "groups" as the keys for these because sets of headers arent valid as keys in python
    '''

    header_to_archetype_dict=dict()
    for header in total_headers:
        #we make sets in each iteration to preserve the order of the list
        header_to_archetype_dict[header]=FORM_HEADER_DICT_REVERSE[header].intersection(set(selected_archetypes))
    
    #now we have {header:{archetypes}} for each header
    #we want to group these sets (we want the reverse of this {{archetypes}:header})
    #but we cannot use sets as keys
    #so we settle for a scheme
    #{group_number:[headers]}
    #combined with
    #{group_number:[archetypes]}
    #first we get the group number keys
    #this is a list of archtypes [archetypes]
    unique_groups=[]
    for header in header_to_archetype_dict.keys():
        #if the archetype set is not in unique groups
        if header_to_archetype_dict[header] not in unique_groups:
            unique_groups.append(header_to_archetype_dict[header])

    #now for each unique group, get the list of headers associated
    group_to_header_dict={i:list() for i in range(len(unique_groups))}
    for temp_header in header_to_archetype_dict:
        temp_group_number=unique_groups.index(header_to_archetype_dict[temp_header])
        group_to_header_dict[temp_group_number].append(temp_header)
    #{group_number:[archetypes]} is the same thing as the unique groups list
    group_to_archetype_dict={i:element for i,element in enumerate(unique_groups)}

    return group_to_header_dict,group_to_archetype_dict

def update_excel_sheet_sample_formatting(workbook,worksheet,group_to_header_dict,group_to_archetype_dict):

    merge_format_dict=dict()
    for group_id in group_to_header_dict.keys():
        merge_format_dict[group_id]=workbook.add_format(
            {
                'align': 'center',
                'valign': 'vcenter',
                'fg_color': COLOR_LIST[group_id]
            }
        )
    start_cell=0
    for group_id in merge_format_dict.keys():
        end_cell=start_cell+len(group_to_header_dict[group_id])-1
        if (end_cell-start_cell)>=1:
            start_cell_xl=xl_rowcol_to_cell(0,start_cell)
            end_cell_xl=xl_rowcol_to_cell(0,end_cell)
            start_cell=start_cell+len(group_to_header_dict[group_id])
            merge_string='Associated with: '+(', '.join(group_to_archetype_dict[group_id]))
            worksheet.merge_range(
                start_cell_xl+':'+end_cell_xl,
                merge_string,
                merge_format_dict[group_id]
            )
        elif (end_cell-start_cell)==0:
            start_cell_xl=xl_rowcol_to_cell(0,start_cell)
            end_cell_xl=xl_rowcol_to_cell(0,end_cell)
            start_cell=start_cell+len(group_to_header_dict[group_id])
            non_merge_string='Associated with: '+(', '.join(group_to_archetype_dict[group_id]))
            worksheet.write(start_cell_xl,non_merge_string,merge_format_dict[group_id])
    worksheet.autofit()

    return workbook, worksheet

def fill_title_sheet(temp_writer,workbook,worksheet):
    worksheet=temp_writer.sheets['title_page']
    worksheet.hide_gridlines()
    top_format=workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'font_size':16
    })
    rule_format=workbook.add_format({
        #'bold': 1,
        #'border': 1,
        'align': 'left',
        'valign': 'vcenter',
        'font_size':16
    })
        
    #write the #first sheet
    worksheet.merge_range('B2:L2','Guidelines',top_format)
    worksheet.merge_range('C4:S4','One Sample Per Row',rule_format)
    worksheet.merge_range('C6:S6','Use fragments/phrases not descriptions ("Mediterranean Diet" not "assorted fish, whole grains, plant oils, etc."',rule_format)
    worksheet.merge_range('C8:S8','Leave unused columns empty',rule_format)         
    worksheet.merge_range('C10:S10','Insert a column for multiple values of the same time. e.g., multiple drugs',rule_format)    

    return workbook, worksheet

@callback(
    [
        Output(component_id="download_form", component_property="data"),
    ],
    [
        Input(component_id='button_form', component_property='n_clicks'),
    ],
    [
        State(component_id='sample_checklist', component_property='value'),
        State(component_id='study_checklist',component_property='value'),
    ],
)
def generate_form(button_form_n_clicks,sample_checklist_options,study_checklist_options):
    '''
    creates the form that is downloaded by users
    '''

    #a potential improvement would be to generate a visible error if nothing is checked
    if sample_checklist_options==None and study_checklist_options==None:
        raise PreventUpdate
    
    if sample_checklist_options==None:
        sample_checklist_options=[]
    if study_checklist_options==None:
        study_checklist_options=[]
    
    #multipele archetypes can have the same headers (eg tissue, cells both have species)
    #we want a non-repeating, ordered list of those headers
    total_headers=generate_form_headers(sample_checklist_options+study_checklist_options)

    #get the dicts that define the colors for the excel file
    group_to_header_dict,group_to_archetype_dict=generate_header_colors(sample_checklist_options+study_checklist_options,total_headers)

    #need to rearrange columns to match group order
    column_order_list=sum(group_to_header_dict.values(),[])

    #empty df for excel file
    temp_dataframe=pd.DataFrame.from_dict(
        {element:[] for element in column_order_list}
    )

    #we write to bytes because it is much more versatile
    output_stream=io.BytesIO()
    temp_writer=pd.ExcelWriter(output_stream,engine='xlsxwriter')

    empty_df=pd.DataFrame()
    empty_df.to_excel(temp_writer,sheet_name='title_page',index=False)
    temp_dataframe.to_excel(temp_writer,sheet_name='sample_sheet',index=False,startrow=1)

    #https://xlsxwriter.readthedocs.io/working_with_pandas.html
    #https://community.plotly.com/t/generate-multiple-tabs-in-excel-file-with-dcc-send-data-frame/53460/7
    workbook=temp_writer.book
    worksheet=temp_writer.sheets['sample_sheet']

    #for each group in group_to_header_dict,group_to_archetype_dict
    #ascertain the number of involved columns in the group
    #ascertain the number of already seen columns
    #merge ((number of involved columns) offset by (number of already seen columns))
    #write text and color ((number of involved columns) offset by (number of already seen columns))
    #for each group, make a format
    #write and color the curation sheet
    
    workbook, worksheet=update_excel_sheet_sample_formatting(workbook,worksheet,group_to_header_dict,group_to_archetype_dict)

    workbook, worksheet=fill_title_sheet(temp_writer,workbook,worksheet)


    temp_writer.save()
    temp_data=output_stream.getvalue()

    return [
        dcc.send_bytes(temp_data,"binbase_sample_ingestion_form.xlsx")
    ]


def generate_curated_colors(worksheet,dataframe):
    '''

    '''

    print('inside_curated_colors')
    print(dataframe)

    max_col=worksheet.max_column

    #find the furthest left archetype that has text
    for temp_col in worksheet.iter_cols(min_row=1,max_row=1,max_col=max_col):
        #print(temp_col[0].fill)
        #print('')
        if temp_col[0].internal_value != None:
            current_internal_text=temp_col[0].internal_value
            break
            
    group_to_header_dict=dict()
    group_to_text_dict=dict()        
    current_group=0
    group_to_header_dict[current_group]=set()
    group_to_text_dict[current_group]=current_internal_text

    #merged_cells=worksheet.merged_cells.ranges
    #print(merged_cells)
    for temp_col in worksheet.iter_cols(min_row=1,max_row=2,max_col=max_col):
    #     if (temp_col[0].coordinate in merged_cells):
    #         group_to_header_dict[current_group].append(
    #             temp_col[1].internal_value
    #         )    
        try:
            if (temp_col[0].internal_value == current_internal_text) or (temp_col[0].internal_value is None):
                group_to_header_dict[current_group].add(
                    temp_col[1].internal_value
                )
            else:
                current_group+=1
                current_internal_text=temp_col[0].internal_value
                group_to_text_dict[current_group]=current_internal_text
                group_to_header_dict[current_group]={temp_col[1].internal_value}
        except AttributeError:
            group_to_header_dict[current_group].add(
                temp_col[1].internal_value
            )

    #[group_to_header_dict[item]:=list(group_to_header_dict[item]) for item in group_to_header_dict.keys()]
    #convert to list so that we can use the dash object store, which has json rules
    for item in group_to_header_dict.keys():
        group_to_header_dict[item]=list(group_to_header_dict[item])

    return group_to_header_dict,group_to_text_dict


@callback(
    [
        Output(component_id="upload_form",component_property="children"),
        Output(component_id="main_store",component_property="data")
    ],
    [
        Input(component_id="upload_form", component_property="contents"),
    ],
    [
        State(component_id="upload_form", component_property="filename"),
        #State(component_id="upload_form", component_property="last_modified"),
        #State(component_id="main_store",component_property="data"),
    ],
    prevent_initial_call=True
)
def upload_form(
    upload_form_contents,
    upload_form_filename,
    #upload_form_last_modified,
    #main_store_data
):
    if upload_form_contents==None:
        raise PreventUpdate
    
    '''
    accept the form back from the user

    need to have a more fully fledged format-checking and error throwing suite
    '''

    content_type, content_string = upload_form_contents.split(',')

    decoded=base64.b64decode(content_string)

    workbook=openpyxl.load_workbook(io.BytesIO(decoded))
    worksheet=workbook['sample_sheet']
    

    temp_dataframe=pd.read_excel(
        io.BytesIO(decoded),
        sheet_name='sample_sheet',
        skiprows=1
        #index_col=False
    )
    print(temp_dataframe)
    print('what if we dont skip first row')

    temp_dataframe_2=pd.read_excel(
        io.BytesIO(decoded),
        sheet_name='sample_sheet',
        header=None,
        nrows=1
    )
    print(temp_dataframe_2)
    #need to set the temp_dataframe_2 headers to be those from the temp_dataframe in order for the concat to work
    temp_dataframe_2.columns=temp_dataframe.columns

    temp_dataframe=pd.concat(
        [temp_dataframe,temp_dataframe_2],
        axis='index',
        ignore_index=True
    )
    print(temp_dataframe)

    temp_dataframe=split_columns_if_delimited(temp_dataframe)#,temp_dataframe_2)

    #group_to_header_dict_curated,group_to_text_dict_curated=generate_curated_colors(worksheet,temp_dataframe)


    temp_dataframe_as_json=temp_dataframe.to_json(orient='records')

    displayed_name=html.Div([upload_form_filename],className='text-center')

    
    print(temp_dataframe)

    store_dict={
        'input_dataframe':temp_dataframe_as_json,
        #'group_to_header_dict_curated':group_to_header_dict_curated,
        #'group_to_text_dict_curated':group_to_text_dict_curated
    }

    #pprint(store_dict)

    return [displayed_name,store_dict]